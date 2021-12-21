import datetime
import http.client
import json
import time

from openpyxl import Workbook

x = '{"Temperature_GH":4,"Humidity_GH":3,"Temperature_TH":2,"Humidity_TH":1,"Temperature_FH":88,"Humidity_FH":99,"DS18B20_TOP":2,"DS18B20_BOTTOM":5,"case_state":2,"mqtt_connected":true}'


class TestRead:
    def __init__(self,ip):
        self.interval = input("Time between readings (s):")
        self.testState = []
        self.time = []
        self.data = 0
        while True:

            self.testState.append(input("1.fanin\n\r2.fanout\n\r3.pump\n\r4.sprinkler\n\r"
                                   "put in the the components that should be ON! Example: 13\n\r"
                                   "if no more components needs to be selected press enter\n\r"))
            if not self.testState[-1]:
                break
            self.time.append(input("For how long do the components needs to be on (s):"))

        self.connect = http.client.HTTPConnection(ip)

    def get_data(self,ws):
        start = time.monotonic()
        for B in range(2, int(self.time[-1])):

            time.sleep(0.1)
            #self.connect.request("GET", "/api/Stats")
            #response = self.connect.getresponse()
           # print("Status: {} and reason: {}".format(response.status, response.reason))
            #body = response.read()
            self.data = json.loads(x)
            ws.cell(B, 1, time.monotonic() - start)
            ws.cell(B, 2, self.data["Temperature_GH"])
            ws.cell(B, 3, self.data["Humidity_GH"])
            ws.cell(B, 4, self.data["Temperature_TH"])
            ws.cell(B, 5, self.data["Humidity_TH"])
            ws.cell(B, 6, self.data["Temperature_FH"])
            ws.cell(B, 7, self.data["Humidity_FH"])
            ws.cell(B, 8, self.data["DS18B20_TOP"])
            ws.cell(B, 9, self.data["DS18B20_BOTTOM"])



if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.title = "measurements"
    ws.cell(1, 1, "Time")
    ws.cell(1, 2, "Temperature GH")
    ws.cell(1, 3, "Humidity GH")
    ws.cell(1, 4, "Temperature FH")
    ws.cell(1, 5, "Humidity FH")
    ws.cell(1, 6, "Temperature TH")
    ws.cell(1, 7, "Humidity TH")
    ws.cell(1, 8, "DS18B20 TOP")
    ws.cell(1, 9, "DS18B20 BOTTOM")
    tester = TestRead("192.168.178.1")
    tester.get_data(ws)
    wb.save("sample.xlsx")